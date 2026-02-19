"""
TIA Tag Import/Export - CSV/Excel conversion for TIA Portal tags and DB variables.

Export (from PLF, no TIA Portal needed):
  - Extract block interface variables (Input/Output/InOut/Static/Temp)
  - Export to CSV or Excel (.xlsx)

Import (generates TIA Openness XML):
  - Read tags from CSV or Excel
  - Generate PLC tag table XML (for %I/%Q/%M addresses)
  - Generate Data Block XML (for DB variables)

Usage:
    # Export from TIA project
    exporter = TiaTagExporter("path/to/project")
    exporter.export_csv("variables.csv")
    exporter.export_excel("variables.xlsx")  # requires openpyxl

    # Import from CSV to XML
    importer = TiaTagImporter()
    importer.import_csv("my_tags.csv")
    importer.generate_tag_table_xml("IO_Tags.xml")
    importer.generate_db_xml("DB_Process.xml", db_number=10)
"""

import csv
import re
import zlib
import xml.etree.ElementTree as ET
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional


# ─── RID → Data Type Mapping ─────────────────────────────────────────────────

RID_TYPE_MAP = {
    "0x02000001": "Bool",
    "0x02000002": "Byte",
    "0x02000003": "Char",
    "0x02000004": "Int",
    "0x02000005": "Word",
    "0x02000006": "DWord",
    "0x02000007": "DInt",
    "0x02000008": "Real",
    "0x02000009": "Date",
    "0x0200000a": "Time_Of_Day",
    "0x0200000b": "Time",
    "0x0200000c": "S5Time",
    "0x02000010": "Date_And_Time",
    "0x02000014": "String",
    "0x02000019": "Pointer",
    "0x0200001a": "Any",
    "0x02000030": "LReal",
    "0x02000031": "ULInt",
    "0x02000032": "LInt",
    "0x02000033": "LWord",
    "0x02000034": "USInt",
    "0x02000035": "UInt",
    "0x02000036": "UDInt",
    "0x02000037": "SInt",
    "0x02000038": "WChar",
    "0x02000039": "WString",
    "0x0200003a": "LTime",
    "0x02000040": "LDT",
    "0x02000041": "DTL",
    "0x02000043": "DTL",
    "0x02000090": "HW_IO",
    "0x02000091": "HW_DEVICE",
    "0x02000092": "HW_DPMASTER",
    "0x02000093": "HW_DPSLAVE",
    "0x02000094": "HW_IOSYSTEM",
    "0x02000095": "HW_SUBMODULE",
    "0x02000096": "Conn_Any",
    "0x020000a0": "DB_ANY",
    "0x020000a1": "DB_WWW",
    "0x020000a2": "DB_DYN",
    "0x02080022": "F_SYSINFO",  # Safety system info UDT
}


# ─── Data Classes ─────────────────────────────────────────────────────────────

@dataclass
class TagEntry:
    """Represents a single tag or variable."""
    block_name: str = ""
    section: str = ""          # Input, Output, InOut, Static, Temp, Constant
    name: str = ""
    data_type: str = ""
    address: str = ""          # %I0.0, %Q0.0, %MW10, etc. (for PLC tags)
    offset: int = -1           # Byte offset within block (for DB vars)
    initial_value: str = ""
    comment: str = ""
    member_id: int = 0
    rid: str = ""
    group: str = ""            # Subgroup/folder name

    def to_dict(self) -> dict:
        return {
            "Block": self.block_name,
            "Section": self.section,
            "Name": self.name,
            "DataType": self.data_type,
            "Address": self.address,
            "Offset": self.offset if self.offset >= 0 else "",
            "InitialValue": self.initial_value,
            "Comment": self.comment,
            "Group": self.group,
        }


CSV_HEADER = ["Block", "Section", "Name", "DataType", "Address", "Offset", "InitialValue", "Comment", "Group"]


# ─── Exporter ─────────────────────────────────────────────────────────────────

class TiaTagExporter:
    """Exports tags and variables from a TIA Portal project to CSV/Excel."""

    def __init__(self, project_path: str):
        self.project_path = Path(project_path)
        if self.project_path.suffix in (".ap14", ".ap15", ".ap16", ".ap17", ".ap18", ".ap19", ".ap20"):
            self.project_path = self.project_path.parent
        self._tags: list[TagEntry] = []
        self._parsed = False

    def parse(self) -> list[TagEntry]:
        """Parse all tags/variables from the project. Returns list of TagEntry."""
        if self._parsed:
            return self._tags

        plf_path = self.project_path / "System" / "PEData.plf"
        if not plf_path.exists():
            raise FileNotFoundError(f"PEData.plf not found: {plf_path}")

        data = plf_path.read_bytes()
        blocks = self._extract_zlib_blocks(data)

        # Parse block interfaces from XML data pages
        self._parse_block_interfaces(blocks)

        self._parsed = True
        return self._tags

    def export_csv(self, filepath: str, delimiter: str = ";"):
        """
        Export all tags to CSV.

        Args:
            filepath: Output CSV file path
            delimiter: CSV delimiter (default ";" for German Excel compatibility)
        """
        if not self._parsed:
            self.parse()

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_HEADER, delimiter=delimiter)
            writer.writeheader()
            for tag in self._tags:
                writer.writerow(tag.to_dict())

        print(f"Exported {len(self._tags)} tags to {path}")

    def export_excel(self, filepath: str):
        """
        Export all tags to Excel (.xlsx). Requires openpyxl.

        Args:
            filepath: Output .xlsx file path
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export: pip install openpyxl")

        if not self._parsed:
            self.parse()

        wb = Workbook()

        # Group tags by block name
        blocks: dict[str, list[TagEntry]] = {}
        for tag in self._tags:
            key = tag.block_name or "Unassigned"
            blocks.setdefault(key, []).append(tag)

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_fills = {
            "Input": PatternFill(start_color="E2EFDA", fill_type="solid"),
            "Output": PatternFill(start_color="FCE4D6", fill_type="solid"),
            "InOut": PatternFill(start_color="DDEBF7", fill_type="solid"),
            "Static": PatternFill(start_color="FFF2CC", fill_type="solid"),
            "Temp": PatternFill(start_color="F2F2F2", fill_type="solid"),
        }
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        first_sheet = True
        for block_name, tags in blocks.items():
            # Sheet name: max 31 chars, no special chars
            sheet_name = re.sub(r'[\\/*?:\[\]]', '_', block_name)[:31]
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(sheet_name)

            # Header row
            for col, header in enumerate(CSV_HEADER, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Data rows
            for row_idx, tag in enumerate(tags, 2):
                values = list(tag.to_dict().values())
                section = tag.section
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border
                    if section in section_fills:
                        cell.fill = section_fills[section]

            # Auto-fit column widths
            for col in range(1, len(CSV_HEADER) + 1):
                max_len = len(CSV_HEADER[col - 1])
                for row in range(2, len(tags) + 2):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[chr(64 + col)].width = min(max_len + 2, 50)

            # Freeze header row
            ws.freeze_panes = "A2"
            # Auto-filter
            ws.auto_filter.ref = f"A1:{chr(64 + len(CSV_HEADER))}{len(tags) + 1}"

        # Overview sheet
        ws_overview = wb.create_sheet("Overview", 0)
        ws_overview.cell(row=1, column=1, value="Block").font = Font(bold=True)
        ws_overview.cell(row=1, column=2, value="Variables").font = Font(bold=True)
        ws_overview.cell(row=1, column=3, value="Input").font = Font(bold=True)
        ws_overview.cell(row=1, column=4, value="Output").font = Font(bold=True)
        ws_overview.cell(row=1, column=5, value="Static").font = Font(bold=True)
        for row_idx, (bname, btags) in enumerate(blocks.items(), 2):
            ws_overview.cell(row=row_idx, column=1, value=bname)
            ws_overview.cell(row=row_idx, column=2, value=len(btags))
            ws_overview.cell(row=row_idx, column=3, value=sum(1 for t in btags if t.section == "Input"))
            ws_overview.cell(row=row_idx, column=4, value=sum(1 for t in btags if t.section == "Output"))
            ws_overview.cell(row=row_idx, column=5, value=sum(1 for t in btags if t.section == "Static"))

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        print(f"Exported {len(self._tags)} tags to {path} ({len(blocks)} blocks)")

    # ── Internal Parsing ──────────────────────────────────────────────────────

    def _extract_zlib_blocks(self, data: bytes) -> list:
        """Find and decompress all zlib blocks."""
        blocks = []
        i = 0
        while i < len(data) - 2:
            if data[i] == 0x78 and data[i + 1] in (0x01, 0x5E, 0x9C, 0xDA):
                try:
                    obj = zlib.decompressobj()
                    dec = obj.decompress(data[i:])
                    comp_sz = len(data[i:]) - len(obj.unused_data)
                    if len(dec) >= 10:
                        blocks.append({"offset": i, "data": dec, "size": len(dec)})
                        i += comp_sz
                        continue
                except Exception:
                    pass
            i += 1
        return blocks

    def _parse_block_interfaces(self, blocks: list):
        """Extract all block interface members from XML data pages."""
        # Track which blocks we've seen to avoid duplicates
        seen_signatures = set()

        for blk in blocks:
            if blk["size"] != 4096:
                continue
            data = blk["data"]
            if b"<Member " not in data or b"<Root RIdSlots" not in data and b"<Member><Member" not in data:
                continue

            text = data.decode("utf-8", errors="replace")

            # Determine block name from context
            block_name = self._identify_block(data, text)

            # Parse sections and members
            members = self._parse_members_xml(text, block_name)

            # Deduplicate: use a signature based on member names
            if members:
                sig = f"{block_name}:" + ",".join(m.name for m in members[:5])
                if sig in seen_signatures:
                    continue
                seen_signatures.add(sig)
                self._tags.extend(members)

    def _identify_block(self, data: bytes, text: str) -> str:
        """Try to identify the block name from context clues."""
        # Safety blocks
        if b"F_PROG_DAT" in data or b"F_RTG_DAT" in data:
            if b"F_SYSINFO" in data:
                return "F_CTRL_1 (Safety FB)"
            return "SafeSys_DB (F-System)"
        if b"_dnVKE_" in data or b"_lnCACHE" in data or b"_dnACT_MOD_MODE_F_B" in data:
            return "F_CTRL_Runtime (Safety Internal)"
        if b"ChannelInfo" in data and b"ALID" in data:
            return "OB82_DiagAlarm"
        if b"DB_INIT_C" in data and b"MAX_CYC_L" in data:
            return "F_CTRL_DB (Safety Data)"

        # Generic: look for external type references
        ext_match = re.search(r'ExternalType[^>]*Name="([^"]+)"', text)
        if ext_match:
            return f"Block_ref_{ext_match.group(1)}"

        # Look at member names for hints
        first_members = re.findall(r'Member[^>]*Name="([^"]+)"', text)[:3]
        if first_members:
            joined = "_".join(first_members[:2])
            if len(joined) > 30:
                joined = joined[:30]
            return f"Block_{joined}"

        return f"Block_0x{0:04x}"

    def _parse_members_xml(self, text: str, block_name: str) -> list[TagEntry]:
        """Parse <Member> elements from XML text."""
        members = []

        # Determine current section context
        # Sections are defined by: <Member ID="2" Name="Input" SubPartIndex="0" />
        section_map = {}
        for m in re.finditer(
            r'<Member\s+ID="(\d+)"\s+Name="(Input|Output|InOut|Static|Temp|Constant)"'
            r'\s+SubPartIndex="(\d+)"',
            text,
        ):
            section_map[m.group(1)] = m.group(2)

        # Parse actual data members (they have RID attribute)
        current_section = "Static"  # Default section

        # Check if there's a section structure
        # Members within a section parent have the section context
        for section_name in ["Input", "Output", "InOut", "Static", "Temp", "Constant"]:
            # Find section opening: Name="Static" ... n="130" ...>
            pattern = rf'Name="{section_name}"[^>]*>'
            section_match = re.search(pattern, text)
            if not section_match:
                continue

            # Find members that appear after this section marker
            section_start = section_match.end()

            # Find the next section or end
            next_section = len(text)
            for other_section in ["Input", "Output", "InOut", "Static", "Temp", "Constant"]:
                if other_section == section_name:
                    continue
                other_pattern = rf'Name="{other_section}"[^>]*SubPartIndex'
                other_match = re.search(other_pattern, text[section_start:])
                if other_match:
                    next_section = min(next_section, section_start + other_match.start())

            section_text = text[section_start:next_section]

            # Extract members with RID (actual variables, not section headers)
            for m in re.finditer(
                r'<Member\s+ID="(\d+)"\s+Name="([^"]+)"\s+RID="([^"]+)"'
                r'(?:\s+Type="([^"]+)")?'
                r'(?:\s+SubPartIndex="[^"]*")?'
                r'(?:\s+StdO="(\d+)")?'
                r'(?:\s+LID="(\d+)")?',
                section_text,
            ):
                rid = m.group(3).lower()
                data_type = m.group(4) if m.group(4) else RID_TYPE_MAP.get(rid, rid)
                offset = int(m.group(5)) if m.group(5) else -1

                members.append(TagEntry(
                    block_name=block_name,
                    section=section_name,
                    name=m.group(2),
                    data_type=data_type,
                    offset=offset,
                    member_id=int(m.group(1)),
                    rid=rid,
                ))

        # If no section structure found, parse all members as flat list
        if not members:
            for m in re.finditer(
                r'<Member\s+ID="(\d+)"\s+Name="([^"]+)"\s+RID="([^"]+)"'
                r'(?:\s+Type="([^"]+)")?'
                r'(?:\s+SubPartIndex="[^"]*")?'
                r'(?:\s+StdO="(\d+)")?'
                r'(?:\s+LID="(\d+)")?',
                text,
            ):
                rid = m.group(3).lower()
                data_type = m.group(4) if m.group(4) else RID_TYPE_MAP.get(rid, rid)
                offset = int(m.group(5)) if m.group(5) else -1
                name = m.group(2)

                # Skip section headers
                if name in ("Input", "Output", "InOut", "Static", "Temp", "Constant", "Return"):
                    continue

                members.append(TagEntry(
                    block_name=block_name,
                    section="Static",
                    name=name,
                    data_type=data_type,
                    offset=offset,
                    member_id=int(m.group(1)),
                    rid=rid,
                ))

        return members


# ─── Importer ─────────────────────────────────────────────────────────────────

class TiaTagImporter:
    """Imports tags from CSV/Excel and generates TIA Openness XML."""

    def __init__(self):
        self._tags: list[TagEntry] = []

    @property
    def tags(self) -> list[TagEntry]:
        return self._tags

    def import_csv(self, filepath: str, delimiter: str = ";"):
        """
        Import tags from a CSV file.

        Expected columns: Block, Section, Name, DataType, Address, Offset, InitialValue, Comment, Group
        Minimum required: Name, DataType
        """
        path = Path(filepath)
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                tag = TagEntry(
                    block_name=row.get("Block", ""),
                    section=row.get("Section", "Static"),
                    name=row.get("Name", ""),
                    data_type=row.get("DataType", ""),
                    address=row.get("Address", ""),
                    offset=int(row["Offset"]) if row.get("Offset") and row["Offset"] != "" else -1,
                    initial_value=row.get("InitialValue", ""),
                    comment=row.get("Comment", ""),
                    group=row.get("Group", ""),
                )
                if tag.name and tag.data_type:
                    self._tags.append(tag)

        print(f"Imported {len(self._tags)} tags from {path}")

    def import_excel(self, filepath: str, sheet_name: str = None):
        """
        Import tags from an Excel file. Requires openpyxl.

        Args:
            filepath: .xlsx file path
            sheet_name: Specific sheet to import (None = all sheets except 'Overview')
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel import: pip install openpyxl")

        wb = load_workbook(filepath, read_only=True)
        sheets = [sheet_name] if sheet_name else [s for s in wb.sheetnames if s != "Overview"]

        for sname in sheets:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # First row is header
            header = [str(h) if h else "" for h in rows[0]]
            for row in rows[1:]:
                row_dict = {header[i]: (row[i] if i < len(row) else "") for i in range(len(header))}
                tag = TagEntry(
                    block_name=str(row_dict.get("Block", sname) or sname),
                    section=str(row_dict.get("Section", "Static") or "Static"),
                    name=str(row_dict.get("Name", "") or ""),
                    data_type=str(row_dict.get("DataType", "") or ""),
                    address=str(row_dict.get("Address", "") or ""),
                    offset=int(row_dict["Offset"]) if row_dict.get("Offset") and str(row_dict["Offset"]).strip() else -1,
                    initial_value=str(row_dict.get("InitialValue", "") or ""),
                    comment=str(row_dict.get("Comment", "") or ""),
                    group=str(row_dict.get("Group", "") or ""),
                )
                if tag.name and tag.data_type:
                    self._tags.append(tag)

        wb.close()
        print(f"Imported {len(self._tags)} tags from {filepath}")

    def add_tag(self, name: str, data_type: str, address: str = "", comment: str = "",
                section: str = "Static", block_name: str = "", initial_value: str = ""):
        """Programmatically add a single tag."""
        self._tags.append(TagEntry(
            block_name=block_name, section=section, name=name,
            data_type=data_type, address=address, comment=comment,
            initial_value=initial_value,
        ))

    def generate_tag_table_xml(self, filepath: str, table_name: str = "Imported_Tags",
                               language: str = "de-DE"):
        """
        Generate a TIA Openness PLC Tag Table XML from imported tags.
        Only includes tags that have an address (%I, %Q, %M, etc.).

        Args:
            filepath: Output XML path
            table_name: PLC tag table name
            language: Comment language (de-DE, en-US, etc.)
        """
        io_tags = [t for t in self._tags if t.address and t.address.startswith("%")]
        if not io_tags:
            io_tags = self._tags  # Fallback: use all tags

        doc = ET.Element("Document")
        ET.SubElement(doc, "Engineering", version="V14")

        tag_table = ET.SubElement(doc, "SW.Tags.PlcTagTable", ID="0")
        attr_list = ET.SubElement(tag_table, "AttributeList")
        name_elem = ET.SubElement(attr_list, "Name")
        name_elem.text = table_name

        obj_list = ET.SubElement(tag_table, "ObjectList")

        for idx, tag in enumerate(io_tags):
            plc_tag = ET.SubElement(obj_list, "SW.Tags.PlcTag", ID=str(idx + 1))
            tag_attrs = ET.SubElement(plc_tag, "AttributeList")

            n = ET.SubElement(tag_attrs, "Name")
            n.text = tag.name

            dt = ET.SubElement(tag_attrs, "DataTypeName")
            dt.text = tag.data_type

            if tag.address:
                addr = ET.SubElement(tag_attrs, "LogicalAddress")
                addr.text = tag.address

            if tag.comment:
                comment_elem = ET.SubElement(tag_attrs, "Comment")
                ml = ET.SubElement(comment_elem, "MultiLanguageText", Lang=language)
                ml.text = tag.comment

        self._save_xml(doc, filepath)
        print(f"Generated tag table XML: {filepath} ({len(io_tags)} tags)")

    def generate_db_xml(self, filepath: str, db_name: str = "DB_Imported",
                        db_number: int = 1, optimized: bool = True,
                        language: str = "de-DE"):
        """
        Generate a TIA Openness Global DB XML from imported tags.

        Args:
            filepath: Output XML path
            db_name: Data block name
            db_number: DB number
            optimized: Optimized block access (True for S7-1500)
            language: Comment language
        """
        db_tags = [t for t in self._tags if not t.address or not t.address.startswith("%")]
        if not db_tags:
            db_tags = self._tags

        doc = ET.Element("Document")
        ET.SubElement(doc, "Engineering", version="V14")

        block = ET.SubElement(doc, "SW.Blocks.GlobalDB", ID=str(db_number))
        attr_list = ET.SubElement(block, "AttributeList")

        n = ET.SubElement(attr_list, "Name")
        n.text = db_name
        num = ET.SubElement(attr_list, "Number")
        num.text = str(db_number)

        if not optimized:
            ml = ET.SubElement(attr_list, "MemoryLayout")
            ml.text = "Standard"

        interface = ET.SubElement(attr_list, "Interface")
        sections = ET.SubElement(
            interface, "Sections",
            xmlns="http://www.siemens.com/automation/Openness/SW/Interface/v3",
        )
        static_section = ET.SubElement(sections, "Section", Name="Static")

        for tag in db_tags:
            member = ET.SubElement(static_section, "Member", Name=tag.name, Datatype=tag.data_type)
            if tag.initial_value:
                sv = ET.SubElement(member, "StartValue")
                sv.text = tag.initial_value
            if tag.comment:
                comment_elem = ET.SubElement(member, "Comment")
                ml = ET.SubElement(comment_elem, "MultiLanguageText", Lang=language)
                ml.text = tag.comment

        self._save_xml(doc, filepath)
        print(f"Generated DB XML: {filepath} ({len(db_tags)} variables)")

    def generate_fb_xml(self, filepath: str, fb_name: str = "FB_Imported",
                        fb_number: int = 1, language_code: str = "SCL",
                        comment_lang: str = "de-DE"):
        """
        Generate a TIA Openness FB XML from imported tags (preserving sections).

        Args:
            filepath: Output XML path
            fb_name: Function block name
            fb_number: FB number
            language_code: Programming language (SCL, LAD, FBD)
            comment_lang: Comment language
        """
        doc = ET.Element("Document")
        ET.SubElement(doc, "Engineering", version="V14")

        block = ET.SubElement(doc, "SW.Blocks.FB", ID=str(fb_number))
        attr_list = ET.SubElement(block, "AttributeList")

        n = ET.SubElement(attr_list, "Name")
        n.text = fb_name
        num = ET.SubElement(attr_list, "Number")
        num.text = str(fb_number)
        lang = ET.SubElement(attr_list, "ProgrammingLanguage")
        lang.text = language_code

        interface = ET.SubElement(attr_list, "Interface")
        sections = ET.SubElement(
            interface, "Sections",
            xmlns="http://www.siemens.com/automation/Openness/SW/Interface/v3",
        )

        for section_name in ["Input", "Output", "InOut", "Static", "Temp", "Constant"]:
            section_tags = [t for t in self._tags if t.section == section_name]
            section = ET.SubElement(sections, "Section", Name=section_name)
            for tag in section_tags:
                member = ET.SubElement(section, "Member", Name=tag.name, Datatype=tag.data_type)
                if tag.initial_value:
                    sv = ET.SubElement(member, "StartValue")
                    sv.text = tag.initial_value
                if tag.comment:
                    comment_elem = ET.SubElement(member, "Comment")
                    ml = ET.SubElement(comment_elem, "MultiLanguageText", Lang=comment_lang)
                    ml.text = tag.comment

        self._save_xml(doc, filepath)
        print(f"Generated FB XML: {filepath} ({len(self._tags)} variables)")

    @staticmethod
    def _save_xml(root: ET.Element, filepath: str):
        """Save XML with proper formatting."""
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        tree = ET.ElementTree(root)
        ET.indent(tree, space="  ")
        with open(path, "w", encoding="utf-8") as f:
            f.write('<?xml version="1.0" encoding="utf-8"?>\n')
            tree.write(f, encoding="unicode", xml_declaration=False)


# ─── Convenience Functions ────────────────────────────────────────────────────

def export_project_tags(project_path: str, output: str, format: str = "csv"):
    """
    One-liner: Export all tags from a TIA project to CSV or Excel.

    Args:
        project_path: Path to TIA project folder
        output: Output file path (.csv or .xlsx)
        format: "csv" or "excel"
    """
    exporter = TiaTagExporter(project_path)
    exporter.parse()
    if format == "excel" or output.endswith(".xlsx"):
        exporter.export_excel(output)
    else:
        exporter.export_csv(output)


def csv_to_tag_table(csv_path: str, xml_path: str, table_name: str = "Imported_Tags"):
    """One-liner: Convert CSV to TIA Tag Table XML."""
    imp = TiaTagImporter()
    imp.import_csv(csv_path)
    imp.generate_tag_table_xml(xml_path, table_name)


def csv_to_db(csv_path: str, xml_path: str, db_name: str = "DB_Imported", db_number: int = 1):
    """One-liner: Convert CSV to TIA Data Block XML."""
    imp = TiaTagImporter()
    imp.import_csv(csv_path)
    imp.generate_db_xml(xml_path, db_name, db_number)


# ─── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Usage:")
        print("  Export: python tia_tag_export.py export <project_path> <output.csv|.xlsx>")
        print("  Import: python tia_tag_export.py import <input.csv> <output.xml> [tag_table|db|fb]")
        print()
        print("Examples:")
        print('  python tia_tag_export.py export "D:/Projects/MyProject" tags.csv')
        print('  python tia_tag_export.py export "D:/Projects/MyProject" tags.xlsx')
        print('  python tia_tag_export.py import tags.csv IO_Tags.xml tag_table')
        print('  python tia_tag_export.py import vars.csv DB_Process.xml db')
        sys.exit(1)

    command = sys.argv[1]

    if command == "export":
        project = sys.argv[2]
        output = sys.argv[3]
        export_project_tags(project, output)

    elif command == "import":
        csv_file = sys.argv[2]
        xml_file = sys.argv[3]
        mode = sys.argv[4] if len(sys.argv) > 4 else "tag_table"

        imp = TiaTagImporter()
        if csv_file.endswith(".xlsx"):
            imp.import_excel(csv_file)
        else:
            imp.import_csv(csv_file)

        if mode == "db":
            imp.generate_db_xml(xml_file)
        elif mode == "fb":
            imp.generate_fb_xml(xml_file)
        else:
            imp.generate_tag_table_xml(xml_file)
    else:
        print(f"Unknown command: {command}")
        sys.exit(1)
